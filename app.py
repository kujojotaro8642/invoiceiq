"""
InvoiceIQ - Full SaaS Backend
Fixed: persistent DB path, proper session config, smart auth flow
"""

import os, re, json, time, uuid, io
from datetime import datetime, timedelta
from functools import wraps
from flask import Flask, request, jsonify, send_file, session
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import pandas as pd

BASE_DIR      = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = "/tmp/invoiceiq_uploads"
OUTPUT_FOLDER = "/tmp/invoiceiq_outputs"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# ── DB: PostgreSQL on Render, SQLite locally ──────────────────────────────
DATABASE_URL = os.environ.get("DATABASE_URL", "")
if DATABASE_URL.startswith("postgres://"):
    # Render gives postgres:// but SQLAlchemy needs postgresql://
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)
DB_PATH = DATABASE_URL if DATABASE_URL else f"sqlite:////tmp/invoiceiq.db"

app = Flask(__name__, static_folder=BASE_DIR, static_url_path="")

# ── Session & security config ──────────────────────────────────────────────
app.secret_key = os.environ.get("SECRET_KEY", "invoiceiq-secret-2024-change-in-render")
app.config["SQLALCHEMY_DATABASE_URI"]        = DB_PATH
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["UPLOAD_FOLDER"]                  = UPLOAD_FOLDER
app.config["MAX_CONTENT_LENGTH"]             = 16 * 1024 * 1024
app.config["SESSION_COOKIE_SAMESITE"]        = "Lax"
app.config["SESSION_COOKIE_SECURE"]          = False   # set True only with HTTPS + custom domain
app.config["PERMANENT_SESSION_LIFETIME"]     = timedelta(days=30)

CORS(app, supports_credentials=True, origins="*")
db = SQLAlchemy(app)

FREE_LIMIT          = 3        # Free users get 3 file uploads total
PRO_MONTHLY_LIMIT   = 100      # Pro users get 100 file uploads/month
OVERAGE_PRICE       = "₹100"   # Charge per 30 extra files
OVERAGE_FILES       = 30       # Files per overage pack
PAID_PRICE          = "₹499/month"
RAZORPAY_LINK  = os.environ.get("RAZORPAY_LINK", "https://razorpay.me/your-link")
ALLOWED_EXT    = {"jpg", "jpeg", "png", "csv"}

# ── Models ─────────────────────────────────────────────────────────────────
class User(db.Model):
    id             = db.Column(db.Integer, primary_key=True)
    email          = db.Column(db.String(120), unique=True, nullable=False)
    password_hash  = db.Column(db.String(256), nullable=False)
    name           = db.Column(db.String(100), nullable=False)
    is_paid          = db.Column(db.Boolean, default=False)
    pro_expires_at   = db.Column(db.DateTime, nullable=True)   # when Pro expires
    files_used_total = db.Column(db.Integer, default=0)        # total files ever uploaded
    files_used_month = db.Column(db.Integer, default=0)        # files uploaded this month
    overage_files    = db.Column(db.Integer, default=0)        # extra files bought via overage
    month_reset_at   = db.Column(db.DateTime, nullable=True)   # when monthly counter last reset
    created_at       = db.Column(db.DateTime, default=datetime.utcnow)

    def set_password(self, pw):   self.password_hash = generate_password_hash(pw)
    def check_password(self, pw): return check_password_hash(self.password_hash, pw)

    def check_expiry(self):
        """Auto-revoke Pro if subscription has expired."""
        if self.is_paid and self.pro_expires_at and datetime.utcnow() > self.pro_expires_at:
            self.is_paid        = False
            self.pro_expires_at = None
            db.session.commit()

    def reset_month_if_needed(self):
        """Reset monthly file counter on the 1st of each month."""
        now = datetime.utcnow()
        if self.month_reset_at is None or (
            now.year > self.month_reset_at.year or now.month > self.month_reset_at.month
        ):
            self.files_used_month = 0
            self.overage_files    = 0   # overage also resets monthly
            self.month_reset_at   = now
            db.session.commit()

    def days_left(self):
        if not self.is_paid or not self.pro_expires_at:
            return None
        delta = self.pro_expires_at - datetime.utcnow()
        return max(0, delta.days)

    def to_dict(self):
        self.check_expiry()
        self.reset_month_if_needed()
        if self.is_paid:
            total_allowed = PRO_MONTHLY_LIMIT + self.overage_files
            files_left    = max(0, total_allowed - self.files_used_month)
        else:
            total_allowed = FREE_LIMIT
            files_left    = max(0, FREE_LIMIT - self.files_used_total)
        return {
            "id":               self.id,
            "email":            self.email,
            "name":             self.name,
            "is_paid":          self.is_paid,
            "files_used_total": self.files_used_total,
            "files_used_month": self.files_used_month,
            "files_left":       files_left,
            "overage_files":    self.overage_files,
            "pro_expires_at":   self.pro_expires_at.isoformat() if self.pro_expires_at else None,
            "days_left":        self.days_left(),
        }

class PaymentRecord(db.Model):
    """Stores every used Payment ID — prevents reuse by same or different user."""
    id          = db.Column(db.Integer, primary_key=True)
    payment_ref = db.Column(db.String(200), unique=True, nullable=False)
    user_id     = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    used_at     = db.Column(db.DateTime, default=datetime.utcnow)

with app.app_context():
    db.create_all()

# ── Auth helpers ───────────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return jsonify({"error":"Login required","code":"AUTH_REQUIRED"}), 401
        return f(*args, **kwargs)
    return decorated

def get_current_user():
    if "user_id" not in session: return None
    return User.query.get(session["user_id"])

def allowed_file(filename):
    return "." in filename and filename.rsplit(".",1)[1].lower() in ALLOWED_EXT

# ══════════════════════════════════════════════════════════════════════════
# AUTH ROUTES
# ══════════════════════════════════════════════════════════════════════════

@app.route("/auth/check-email", methods=["POST"])
def check_email():
    """Tell frontend whether this email is registered or new."""
    data  = request.get_json()
    email = (data.get("email") or "").strip().lower()
    if not email or not re.match(r"[^@]+@[^@]+\.[^@]+", email):
        return jsonify({"error": "Invalid email"}), 400
    exists = User.query.filter_by(email=email).first() is not None
    return jsonify({"exists": exists, "email": email})

@app.route("/auth/signup", methods=["POST"])
def signup():
    data  = request.get_json()
    name  = (data.get("name") or "").strip()
    email = (data.get("email") or "").strip().lower()
    pw    = data.get("password") or ""
    if not name or not email or not pw:
        return jsonify({"error":"Name, email and password are required"}), 400
    if not re.match(r"[^@]+@[^@]+\.[^@]+", email):
        return jsonify({"error":"Invalid email address"}), 400
    if len(pw) < 6:
        return jsonify({"error":"Password must be at least 6 characters"}), 400
    if User.query.filter_by(email=email).first():
        return jsonify({"error":"An account with this email already exists"}), 409
    user = User(name=name, email=email)
    user.set_password(pw)
    db.session.add(user)
    db.session.commit()
    session.permanent = True
    session["user_id"] = user.id
    return jsonify({"message":"Account created!","user":user.to_dict()}), 201

@app.route("/auth/login", methods=["POST"])
def login():
    data  = request.get_json()
    email = (data.get("email") or "").strip().lower()
    pw    = data.get("password") or ""
    user  = User.query.filter_by(email=email).first()
    if not user or not user.check_password(pw):
        return jsonify({"error":"Invalid email or password"}), 401
    session.permanent = True
    session["user_id"] = user.id
    return jsonify({"message":"Logged in!","user":user.to_dict()})

@app.route("/auth/logout", methods=["POST"])
def logout():
    session.clear()
    return jsonify({"message":"Logged out"})

@app.route("/auth/me")
def me():
    user = get_current_user()
    if not user: return jsonify({"error":"Not logged in","code":"AUTH_REQUIRED"}), 401
    return jsonify({"user":user.to_dict()})

# ── Payment routes ─────────────────────────────────────────────────────────
@app.route("/payment/info")
@login_required
def payment_info():
    return jsonify({"price":PAID_PRICE,"payment_link":RAZORPAY_LINK,"free_limit":FREE_LIMIT,
                    "features":["Unlimited invoice processing","Batch CSV, JPG & PNG","Excel + JSON export","GST auto-detection","Duplicate alerts"]})

@app.route("/payment/activate", methods=["POST"])
@login_required
def activate_paid():
    data = request.get_json()
    ref  = (data.get("payment_ref") or "").strip()

    if not ref:
        return jsonify({"error":"Payment reference required"}), 400

    # Block reuse — check if this Payment ID was already used by anyone
    existing = PaymentRecord.query.filter_by(payment_ref=ref).first()
    if existing:
        return jsonify({"error":"This Payment ID has already been used. Each payment can only activate one account."}), 409

    user = get_current_user()

    # Block if user is already on active Pro (not expired)
    if user.is_paid and user.pro_expires_at and datetime.utcnow() < user.pro_expires_at:
        days = user.days_left()
        return jsonify({"error": f"Your Pro plan is still active for {days} more day(s). You can renew when it expires."}), 400

    # Record the payment ID so it can never be reused
    record = PaymentRecord(payment_ref=ref, user_id=user.id)
    db.session.add(record)

    # Set Pro active for exactly 30 days from now
    user.is_paid        = True
    user.pro_expires_at = datetime.utcnow() + timedelta(days=30)
    user.month_used     = 0
    user.month_reset_at = datetime.utcnow()
    db.session.commit()

    return jsonify({
        "message": f"Pro activated! Valid until {user.pro_expires_at.strftime('%d %b %Y')}.",
        "user": user.to_dict()
    })

@app.route("/payment/overage", methods=["POST"])
@login_required
def activate_overage():
    """User buys 30 extra file uploads for ₹100."""
    data = request.get_json()
    ref  = (data.get("payment_ref") or "").strip()

    if not ref:
        return jsonify({"error": "Payment reference required"}), 400

    user = get_current_user()
    if not user.is_paid:
        return jsonify({"error": "You need an active Pro plan to buy overage packs."}), 400

    # Block reuse of same payment ID
    existing = PaymentRecord.query.filter_by(payment_ref=ref).first()
    if existing:
        return jsonify({"error": "This Payment ID has already been used."}), 409

    record = PaymentRecord(payment_ref=ref, user_id=user.id)
    db.session.add(record)
    user.overage_files += OVERAGE_FILES
    db.session.commit()

    return jsonify({
        "message": f"✅ {OVERAGE_FILES} extra file uploads added!",
        "user":    user.to_dict()
    })


# ── OCR ────────────────────────────────────────────────────────────────────
def extract_text_from_image(filepath):
    try:
        import pytesseract
        from PIL import Image
        return pytesseract.image_to_string(Image.open(filepath).convert("RGB"))
    except Exception:
        return None

def parse_invoice_from_text(text):
    data = {k:None for k in ["invoice_number","date","vendor_name","total_amount","gst_number"]}
    m = re.search(r"(?:invoice\s*(?:no|number|#)[:\s#]*)([\w\-/]+)", text, re.IGNORECASE)
    if m: data["invoice_number"] = m.group(1).strip()
    else:
        m2 = re.search(r"\b(INV[-/]?\d+)\b", text, re.IGNORECASE)
        if m2: data["invoice_number"] = m2.group(1).strip()
    m = re.search(r"\b(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4}|\d{4}[\/\-]\d{2}[\/\-]\d{2}|\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+\d{4})\b", text, re.IGNORECASE)
    if m: data["date"] = m.group(1).strip()
    m = re.search(r"(?:from|vendor|bill\s*from|seller)[:\s]+([A-Za-z0-9 &.,'\-]+)", text, re.IGNORECASE)
    if m: data["vendor_name"] = m.group(1).strip()[:60]
    m = re.search(r"(?:grand\s*total|total\s*amount|total|amount\s*due)[:\s\u20b9$\u20ac\xa3]*([0-9,]+\.?\d*)", text, re.IGNORECASE)
    if m: data["total_amount"] = m.group(1).replace(",","").strip()
    m = re.search(r"\b(\d{2}[A-Z]{5}\d{4}[A-Z][A-Z\d]Z[A-Z\d])\b", text, re.IGNORECASE)
    if m: data["gst_number"] = m.group(1).upper()
    return data

def parse_invoice_from_csv(filepath):
    invoices = []
    ALIASES = {
        "invoice_number":["invoice_number","invoice no","invoice#","inv no","inv_no","invoice id","bill no"],
        "date":["date","invoice_date","bill date","billing date","invoice date"],
        "vendor_name":["vendor_name","vendor","seller","company","from","supplier","party name"],
        "total_amount":["total_amount","total","amount","grand total","amount due","net amount","payable"],
        "gst_number":["gst_number","gst no","gst","gstin","tax id","gst number"],
    }
    try:
        df = pd.read_csv(filepath, dtype=str)
        df.columns = [c.strip().lower() for c in df.columns]
        col_map = {}
        for field, aliases in ALIASES.items():
            for alias in aliases:
                for col in df.columns:
                    if alias in col or col in alias:
                        if col not in col_map: col_map[col] = field
                        break
        for _, row in df.iterrows():
            inv = {k:None for k in ["invoice_number","date","vendor_name","total_amount","gst_number"]}
            for col, field in col_map.items():
                val = row.get(col)
                if pd.notna(val) and str(val).strip(): inv[field] = str(val).strip()
            invoices.append(inv)
    except Exception: pass
    return invoices

def validate_invoice(inv, seen):
    warnings = []
    for f in ["invoice_number","date","vendor_name","total_amount"]:
        if not inv.get(f): warnings.append(f"Missing: {f.replace('_',' ').title()}")
    inv_no = inv.get("invoice_number")
    if inv_no:
        if inv_no in seen: warnings.append(f"Duplicate invoice number: {inv_no}")
        else: seen.add(inv_no)
    return warnings

def export_to_excel(items, output_path):
    rows = []
    for item in items:
        inv = item.get("data") or {}
        rows.append({"File Name":item.get("filename",""),"Invoice Number":inv.get("invoice_number",""),
                     "Date":inv.get("date",""),"Vendor Name":inv.get("vendor_name",""),
                     "Total Amount":inv.get("total_amount",""),"GST Number":inv.get("gst_number",""),
                     "Warnings":"; ".join(item.get("warnings",[])) or "OK"})
    df = pd.DataFrame(rows)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Invoices")
        ws = writer.sheets["Invoices"]
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
        hf = PatternFill("solid", fgColor="1E3A5F")
        ef = PatternFill("solid", fgColor="F8D7DA")
        for ci in range(1, len(df.columns)+1):
            c = ws.cell(row=1, column=ci)
            c.fill = hf; c.font = Font(bold=True, color="FFFFFF", size=11)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25
        for ri, item in enumerate(items, 2):
            for ci in range(1, len(df.columns)+1):
                c = ws.cell(row=ri, column=ci)
                c.alignment = Alignment(vertical="center")
                if item.get("warnings"): c.fill = ef
        for ci, col in enumerate(df.columns, 1):
            vals = [str(ws.cell(row=r, column=ci).value or "") for r in range(1, len(rows)+2)]
            ws.column_dimensions[get_column_letter(ci)].width = min(max((len(v) for v in vals), default=10)+4, 45)
    return output_path

# ── Process route ──────────────────────────────────────────────────────────
@app.route("/process", methods=["POST"])
@login_required
def process_files():
    user = get_current_user()
    files = request.files.getlist("files")
    if not files or all(f.filename=="" for f in files):
        return jsonify({"error":"No files selected"}), 400

    # Auto-check subscription expiry and monthly reset
    user.check_expiry()
    user.reset_month_if_needed()

    # Count valid files in this upload
    valid_files = [f for f in files if f.filename and allowed_file(f.filename)]
    num_files   = len(valid_files)

    if not user.is_paid:
        # Free: count total files ever uploaded
        if user.files_used_total >= FREE_LIMIT:
            return jsonify({"error":"free_limit_reached",
                            "message":f"You've used all {FREE_LIMIT} free file uploads. Upgrade to Pro for 100 files/month.",
                            "payment_link":RAZORPAY_LINK,"price":PAID_PRICE}), 402
    else:
        # Pro: count files this month (base 100 + any overage packs bought)
        total_allowed = PRO_MONTHLY_LIMIT + user.overage_files
        if user.files_used_month >= total_allowed:
            return jsonify({"error":"monthly_limit_reached",
                            "message":f"You've used all {total_allowed} file uploads this month.",
                            "overage_available": True,
                            "overage_price": OVERAGE_PRICE,
                            "overage_files": OVERAGE_FILES,
                            "payment_link":RAZORPAY_LINK}), 402

    seen = set(); results = []; start = time.time(); file_count = 0

    for file in files:
        if not file.filename: continue
        if not allowed_file(file.filename):
            results.append({"filename":file.filename,"status":"error",
                            "error":"Unsupported type. Use JPG, PNG, or CSV.",
                            "data":{},"warnings":[]}); continue

        # Check per-file limit (free users mid-batch)
        if not user.is_paid and (user.files_used_total + file_count) >= FREE_LIMIT:
            results.append({"filename":file.filename,"status":"error",
                            "error":"Free file limit reached. Upgrade to Pro.",
                            "data":{},"warnings":[]}); continue

        # Check Pro monthly limit mid-batch
        if user.is_paid:
            total_allowed = PRO_MONTHLY_LIMIT + user.overage_files
            if (user.files_used_month + file_count) >= total_allowed:
                results.append({"filename":file.filename,"status":"error",
                                "error":"Monthly file limit reached. Buy an overage pack to continue.",
                                "data":{},"warnings":[]}); continue

        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)
        file.save(filepath)
        ext = filename.rsplit(".",1)[1].lower()

        if ext == "csv":
            parsed = parse_invoice_from_csv(filepath)
            if not parsed:
                results.append({"filename":filename,"status":"error",
                                "error":"Could not parse CSV. Check column names.",
                                "data":{},"warnings":[]})
            else:
                # Each CSV file = 1 file upload regardless of rows
                row_results = []
                for idx, inv in enumerate(parsed):
                    w = validate_invoice(inv, seen)
                    row_results.append({
                        "filename": f"{filename} (row {idx+1})",
                        "status":   "warning" if w else "success",
                        "data":     inv,
                        "warnings": w
                    })
                results.extend(row_results)
                file_count += 1   # count the FILE, not the rows
        else:
            raw = extract_text_from_image(filepath)
            if raw is None:
                results.append({"filename":filename,"status":"error",
                                "error":"Image OCR unavailable. Please use CSV format.",
                                "data":{},"warnings":[]})
            else:
                inv = parse_invoice_from_text(raw)
                w   = validate_invoice(inv, seen)
                results.append({"filename":filename,"status":"warning" if w else "success",
                                "data":inv,"warnings":w,"raw_text_preview":raw[:300]})
                file_count += 1   # count the FILE

        try: os.remove(filepath)
        except: pass

    # Update counters — FILE based
    user.files_used_total += file_count
    user.files_used_month += file_count
    db.session.commit()
    elapsed = round(time.time()-start, 2)
    session_id = str(uuid.uuid4())[:8]
    with open(os.path.join(OUTPUT_FOLDER,f"invoices_{session_id}.json"),"w") as f: json.dump(results,f,indent=2)
    export_to_excel(results, os.path.join(OUTPUT_FOLDER,f"invoices_{session_id}.xlsx"))
    if user.is_paid:
        total_allowed   = PRO_MONTHLY_LIMIT + user.overage_files
        files_left      = max(0, total_allowed - user.files_used_month)
    else:
        files_left      = max(0, FREE_LIMIT - user.files_used_total)

    return jsonify({
        "results":          results,
        "processing_time":  elapsed,
        "session_id":       session_id,
        "is_paid":          user.is_paid,
        "files_used_month": user.files_used_month,
        "files_left":       files_left,
        "days_left":        user.days_left(),
        "overage_price":    OVERAGE_PRICE,
        "overage_files":    OVERAGE_FILES,
        "summary": {
            "success":  sum(1 for r in results if r["status"]=="success"),
            "warnings": sum(1 for r in results if r["status"]=="warning"),
            "errors":   sum(1 for r in results if r["status"]=="error"),
        }
    })

@app.route("/download/excel/<session_id>")
@login_required
def download_excel(session_id):
    if not re.match(r'^[a-f0-9]{8}$', session_id): return jsonify({"error":"Invalid session"}), 400
    path = os.path.join(OUTPUT_FOLDER,f"invoices_{session_id}.xlsx")
    if not os.path.exists(path): return jsonify({"error":"File not found"}), 404
    return send_file(path, as_attachment=True, download_name="invoices.xlsx")

@app.route("/download/json/<session_id>")
@login_required
def download_json(session_id):
    if not re.match(r'^[a-f0-9]{8}$', session_id): return jsonify({"error":"Invalid session"}), 400
    path = os.path.join(OUTPUT_FOLDER,f"invoices_{session_id}.json")
    if not os.path.exists(path): return jsonify({"error":"File not found"}), 404
    return send_file(path, as_attachment=True, download_name="invoices.json")

@app.route("/")
def index():
    return send_file(os.path.join(BASE_DIR,"index.html"))

@app.route("/health")
def health():
    return jsonify({"status":"ok","time":datetime.now().isoformat()})

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    print(f"🚀 InvoiceIQ running on http://localhost:{port}")
    app.run(debug=False, host="0.0.0.0", port=port)
